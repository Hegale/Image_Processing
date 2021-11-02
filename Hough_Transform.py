import math
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active

angle = [0, 15, 30, 45, 60, 75, 90, 105, 120, 135, 150, 165] 
rad = [0, 0.262, 0.523, 0.785, 1.047, 1.308, 1.570, 1.832, 2.093, 2.355, 2.617, 2.878] 
H_space = []  #θ, radius 변환값, Hough space 차례로 선언 

#Count를 저장하는 Hough Space 선언
for i in range(31) :  #이때, d = integer r value는 범위를 -15~15로 가정
    l = []           
    for j in range(12) :  
        l.append(0)
    H_space.append(l)

#종료조건이 입력될 때까지, pixel의 x,y좌표 입력
for j in range(100) :
    d = [0]*12
    d_round = [0]*12
    
    x, y = input('X, Y : ').split( )
    x = int(x)
    y = int(y)
    
    if x == -1 and y == -1 :   
        break     #종료조건 : x, y좌표에 -1 입력
    for i in range(12) :
        d[i] = x * math.cos(rad[i]) + y * math.sin(rad[i])
        #d = xcos(θ) + ysin(θ)
        d_round[i] = round(d[i])

    for i in range(12) :
        k = d_round[i]+15
        #d의 범위를 -15~15로 가정했으나, 실제 index는 0~30이므로
        count = H_space[k][i]
        count += 1
        H_space[k][i] = count
        #해당하는 위치에 +=1

for i in range(12) : ws.cell(row = 1, column = i+2).value = angle[i]  

for i in range(31) :
    ws.cell(row = i+2, column = 1).value = i-15
    for j in range(12) :
        ws.cell(row = i+2, column = j+2).value = H_space[i][j]
        if H_space[i][j] > 2 :
            print("(ρ : %d, angle : %d)" %(i-15, angle[j]))

wb.save('result.xlsx')    #파일명 result로 저장
wb.close()
